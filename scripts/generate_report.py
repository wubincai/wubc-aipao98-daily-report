#!/usr/bin/env python3
"""
爱跑98竞赛日报表生成器
生成班组对抗赛+个人销量奖励双表Excel
"""
import argparse, json, os, sys
from collections import defaultdict

DEPENDENCIES = ['xlrd', 'openpyxl']

def check_dependencies():
    missing = []
    for pkg in DEPENDENCIES:
        try:
            __import__(pkg)
        except ImportError:
            missing.append(pkg)
    if missing:
        print(f"[依赖检查] 缺少以下 Python 包: {', '.join(missing)}")
        print(f"请运行: pip install {' '.join(missing)}")
        sys.exit(1)

def parse_time(tstr):
    """从 '2026-07-10 23:49:06' 提取小时:分钟"""
    if not isinstance(tstr, str):
        return None, None
    parts = tstr.split(' ')
    if len(parts) < 2:
        return None, None
    hm = parts[1].split(':')
    if len(hm) < 2:
        return None, None
    return int(hm[0]), int(hm[1])

def in_period(h, m, sh, sm, eh, em):
    """判断时间(h,m)是否在时段[sh:sm, eh:em)内"""
    return (sh * 60 + sm) <= (h * 60 + m) < (eh * 60 + em)

def parse_period_name(name):
    """从时段名称 '06:30-08:00' 解析起止时间"""
    parts = name.split('-')
    sh, sm = parts[0].split(':')
    eh, em = parts[1].split(':')
    return int(sh), int(sm), int(eh), int(em)

def read_data_source(filepath):
    """读取 .xls 交易记录文件"""
    import xlrd
    wb = xlrd.open_workbook(filepath)
    ws = wb.sheet_by_index(0)
    
    # 提取站名 (R6, col 5)
    station_name = None
    for r in range(ws.nrows):
        val1 = str(ws.cell_value(r, 1)).strip()
        val2 = str(ws.cell_value(r, 2)).strip()
        if val1 == '站名:' and val2:
            station_name = val2
            break
    
    # 提取交易数据
    records = []
    for r in range(8, ws.nrows):
        tstr = ws.cell_value(r, 22)
        oil = str(ws.cell_value(r, 4))
        liters_raw = ws.cell_value(r, 9)
        
        if not isinstance(tstr, str) or not tstr.strip():
            continue
        h, m = parse_time(tstr)
        if h is None:
            continue
        try:
            liters = float(liters_raw)
        except (ValueError, TypeError):
            continue
        
        is_diesel = '0号车用柴油' in oil
        is_aipao98 = '爱跑98' in oil
        
        records.append({
            'hour': h, 'minute': m,
            'oil': oil, 'liters': liters,
            'is_diesel': is_diesel,
            'is_aipao98': is_aipao98
        })
    
    return station_name, records

def aggregate_periods(records, periods_config):
    """按配置的时间段汇总数据"""
    result = {}
    for p in periods_config:
        name = p['name']
        sh, sm, eh, em = parse_period_name(name)
        total_liters = 0.0
        aipao98_liters = 0.0
        
        for rec in records:
            if in_period(rec['hour'], rec['minute'], sh, sm, eh, em):
                if not rec['is_diesel']:
                    total_liters += rec['liters']
                if rec['is_aipao98']:
                    aipao98_liters += rec['liters']
        
        result[name] = {
            'total_liters': round(total_liters, 2),
            'aipao98_liters': round(aipao98_liters, 2),
            'teams': p['teams']
        }
    return result

def total_members_in_period(period_config):
    """计算一个时段的总人数"""
    return sum(t['members'] for t in period_config['teams'])

def generate_report(period_data, station_name, date_str, config, output_path):
    """生成双表Excel"""
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    import openpyxl
    
    tf = Font('宋体', 14, bold=True)
    hf = Font('宋体', 12, bold=True, color='FFFFFF')
    df = Font('宋体', 12)
    ca = Alignment(horizontal='center', vertical='center')
    tb = Border(left=Side('thin'), right=Side('thin'), top=Side('thin'), bottom=Side('thin'))
    fi = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Sheet1'
    
    # 计算总行数
    total_rows = 2  # 标题行+表头行
    period_row_info = []  # [(name, start_row, end_row, teams_list), ...]
    
    current_row = 3
    period_names = list(period_data.keys())
    for name in period_names:
        pd_conf = period_data[name]
        teams = pd_conf['teams']
        n_rows = sum(t['members'] for t in teams)
        period_row_info.append((name, current_row, current_row + n_rows - 1, teams))
        total_rows += n_rows
        current_row += n_rows
    
    data_end_row = 2 + sum(sum(t['members'] for t in pd['teams']) for pd in period_data.values())
    
    # 列宽
    col_widths = {'A':14,'B':13,'C':15,'D':13,'E':13,'F':13,'G':19,'H':13,'I':12,'J':16,'K':20,'L':3,'M':14,'N':14,'O':10,'P':16,'R':8}
    for c, w in col_widths.items():
        ws.column_dimensions[c].width = w
    ws.column_dimensions['R'].hidden = True
    
    # 行高
    ws.row_dimensions[1].height = 40
    ws.row_dimensions[2].height = 45
    for r in range(3, data_end_row + 1):
        ws.row_dimensions[r].height = 24
    
    # ===== 左表 =====
    ws.merge_cells('A1:K1')
    ws['A1'] = f'{station_name}爱跑98竞赛日班组对抗赛统计表（{date_str}）'
    ws['A1'].font = tf; ws['A1'].alignment = ca
    
    for i, h in enumerate(['时间段','汽油升数','爱跑98升数','爱跑98占比','班组','组  员',
        '人均爱跑98升数','系数','基础奖励','个人奖励金额','班组奖励金额']):
        c = ws.cell(2, i+1, h)
        c.font = hf; c.alignment = ca; c.border = tb; c.fill = fi
    
    # 合并单元格
    merge_list = ['A1:K1']
    for name, sr, er, teams in period_row_info:
        if sr != er:
            merge_list.append(f'A{sr}:A{er}')
            merge_list.append(f'B{sr}:B{er}')
            merge_list.append(f'C{sr}:C{er}')
            merge_list.append(f'D{sr}:D{er}')
        # E列(班组)和K列(班组奖励)按子组合并
        row_offset = 0
        for ti, t in enumerate(teams):
            gs = sr + row_offset
            ge = gs + t['members'] - 1
            if gs != ge:
                merge_list.append(f'E{gs}:E{ge}')
                merge_list.append(f'K{gs}:K{ge}')
            row_offset += t['members']
    
    for m in merge_list:
        ws.merge_cells(m)
    
    # 数据区统一样式
    for r in range(3, data_end_row + 1):
        for c in range(1, 12):
            ws.cell(r, c).border = tb
            ws.cell(r, c).font = df
            ws.cell(r, c).alignment = ca
    
    # 填充数据
    for name, sr, er, teams in period_row_info:
        d = period_data[name]
        ws.cell(sr, 1).value = name
        ws.cell(sr, 2).value = d['total_liters']
        ws.cell(sr, 2).number_format = '#,##0.00'
        ws.cell(sr, 3).value = d['aipao98_liters']
        ws.cell(sr, 3).number_format = '#,##0.00'
        ws.cell(sr, 4).value = f'=IF(B{sr}>0,C{sr}/B{sr},0)'
        ws.cell(sr, 4).number_format = '0.00%'
        
        # G列(人均爱跑98)
        total_m = total_members_in_period({'teams': teams})
        for r in range(sr, er + 1):
            ws.cell(r, 7).value = f'=IF(COUNTA(F{sr}:F{er})>0,C{sr}/COUNTA(F{sr}:F{er}),0)'
            ws.cell(r, 7).number_format = '#,##0.00'
        
        # E列(班组)和K列(班组奖励)
        row_offset = 0
        for ti, t in enumerate(teams):
            gs = sr + row_offset
            ge = gs + t['members'] - 1
            ws.cell(gs, 5).value = f'=IF(F{gs}<>"",F{gs},"")'
            
            if gs == ge:
                ws.cell(gs, 11).value = f'=IF(AND(H{gs}<>"",I{gs}<>""),G{gs}*H{gs}*I{gs},0)'
            else:
                terms = [f'IF(AND(H{r}<>"",I{r}<>""),G{r}*H{r}*I{r},0)' for r in range(gs, ge+1)]
                ws.cell(gs, 11).value = '=' + '+'.join(terms)
            ws.cell(gs, 11).number_format = '#,##0.00'
            row_offset += t['members']
    
    # I列基础奖励
    base_reward = config.get('base_reward', 0.25)
    for r in range(3, data_end_row + 1):
        ws.cell(r, 9).value = base_reward
        ws.cell(r, 9).number_format = '0.00'
    
    # J列个人奖励金额
    for r in range(3, data_end_row + 1):
        ws.cell(r, 10).value = f'=IF(AND(H{r}<>"",I{r}<>""),G{r}*H{r}*I{r},0)'
        ws.cell(r, 10).number_format = '#,##0.00'
    
    # ===== 右表 =====
    col_m = 13  # M列
    right_start = data_end_row + 1 if data_end_row > 22 else 24
    
    ws.merge_cells(f'M1:P1')
    ws.cell(1, col_m).value = f'{station_name}爱跑98竞赛日个人销量奖励统计表（{date_str}）'
    ws.cell(1, col_m).font = tf; ws.cell(1, col_m).alignment = ca
    
    for i, h in enumerate(['组员','爱跑98升数','系数','个人奖励汇总']):
        c = ws.cell(2, col_m + i, h)
        c.font = hf; c.alignment = ca; c.border = tb; c.fill = fi
    
    # R列辅助(隐藏): 首次出现编号
    ws.cell(2, 18).value = 0
    for r in range(3, data_end_row + 1):
        ws.cell(r, 18).value = f'=IF(F{r}="",N(R{r-1}),IF(COUNTIF(F$3:F{r},F{r})=1,N(R{r-1})+1,N(R{r-1})))'
        ws.cell(r, 18).font = df; ws.cell(r, 18).alignment = ca
    
    # M列紧凑去重
    for r in range(3, right_start + 1):
        ws.cell(r, col_m).value = f'=IFERROR(INDEX(F$3:F${data_end_row},MATCH(ROW()-2,R$3:R${data_end_row},0)),"")'
        ws.cell(r, col_m).font = df; ws.cell(r, col_m).alignment = ca; ws.cell(r, col_m).border = tb
    
    # N-P
    for r in range(3, right_start + 1):
        ws.cell(r, 14).value = f'=IF(M{r}="","",SUMPRODUCT((F$3:F${data_end_row}=M{r})*G$3:G${data_end_row}))'
        ws.cell(r, 15).value = f'=IF(M{r}="","",IFERROR(INDEX(H$3:H${data_end_row},MATCH(M{r},F$3:F${data_end_row},0)),""))'
        ws.cell(r, 16).value = f'=IF(M{r}="","",SUMPRODUCT((F$3:F${data_end_row}=M{r})*J$3:J${data_end_row}))'
        for c in range(14, 17):
            ws.cell(r, c).font = df; ws.cell(r, c).alignment = ca; ws.cell(r, c).border = tb
        ws.cell(r, 14).number_format = '#,##0.00'
        ws.cell(r, 16).number_format = '#,##0.00'
    
    wb.save(output_path)
    return output_path

def main():
    check_dependencies()
    
    parser = argparse.ArgumentParser(description='爱跑98竞赛日报表生成器')
    parser.add_argument('--date', required=True, help='日期，如 7月10日')
    parser.add_argument('--config', required=True, help='配置文件路径')
    parser.add_argument('--data-dir', help='数据源目录（默认 ~/Documents/加油站数据表）')
    parser.add_argument('--output-dir', help='输出目录（默认 ~/Documents/爱跑98竞赛日报表）')
    args = parser.parse_args()
    
    date_str = args.date
    data_dir = args.data_dir or os.path.expanduser('~/Documents/加油站数据表')
    output_dir = args.output_dir or os.path.expanduser('~/Documents/爱跑98竞赛日报表')
    
    # 读取配置
    if not os.path.exists(args.config):
        print(f"[错误] 配置文件不存在: {args.config}")
        print("请先完成时间段、班组、基础奖励的配置")
        sys.exit(1)
    
    with open(args.config, 'r', encoding='utf-8') as f:
        config = json.load(f)
    
    # 查找数据文件
    data_file = os.path.join(data_dir, f'deal_daily_excel.jasper ({date_str}).xls')
    if not os.path.exists(data_file):
        print(f"[错误] 数据文件不存在: {data_file}")
        print(f"请从管控系统下载 {data_file} 后重试")
        sys.exit(1)
    
    # 读取数据源
    print(f"[1/4] 读取数据源: {data_file}")
    station_name, records = read_data_source(data_file)
    if not station_name:
        station_name = "加油站"
    print(f"      站名: {station_name}, 记录数: {len(records)}")
    
    # 按时间段汇总
    print(f"[2/4] 按 {len(config['periods'])} 个时间段汇总...")
    period_data = aggregate_periods(records, config['periods'])
    
    # 创建输出目录
    os.makedirs(output_dir, exist_ok=True)
    output_file = os.path.join(output_dir, f'爱跑98竞赛日班组对抗赛统计表（{date_str}）.xlsx')
    
    # 生成报表
    print(f"[3/4] 生成报表...")
    generate_report(period_data, station_name, date_str, config, output_file)
    
    print(f"[4/4] 完成!")
    print(f"      报表已保存: {output_file}")
    return 0

if __name__ == '__main__':
    sys.exit(main())
